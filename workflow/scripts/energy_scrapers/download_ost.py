#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OST hourly demand scraper

"""
import os
import threading
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed, CancelledError

import yaml
import requests
import pandas as pd
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta
from tqdm import tqdm
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ───────────── Tunables ──────────────────────────────────────────────────── #
BASE_URL        = "https://ost.al/wp-content/uploads"
SUFFIXES        = ("", "-1", "-2", "-3", "-4", "-001", "-002", "-003")
FOLDERS         = (0, 1)   # month folder offsets: current & next
HTTP_TIMEOUT    = 3        # seconds per request
DEFAULT_WORKERS = 32
VERBOSE         = False
FAIL_THRESHOLD  = 2        # how many different days must 404 before we skip
# ─────────────────────────────────────────────────────────────────────────── #

# Thread‑local requests.Session (keep‑alive)
_thread_local = threading.local()

def _session() -> requests.Session:
    if not hasattr(_thread_local, "sess"):
        sess = requests.Session()
        adapter = requests.adapters.HTTPAdapter(pool_connections=8, pool_maxsize=8, max_retries=2)
        sess.mount("https://", adapter)
        sess.mount("http://",  adapter)
        _thread_local.sess = sess
    return _thread_local.sess

# Global caches guarded by a lock
_cache_lock = threading.Lock()
working    : set[tuple[int,int,str]]     = set()          # combos that succeeded once
fail_count : dict[tuple[int,int,str],int] = {}            # 404 counter per combo

# ───────────── Excel helpers ─────────────────────────────────────────────── #
CELL_DATE = "C158"
ROW_START, ROW_END = 160, 184  # inclusive range in sheet column F


def _open_wb(raw: bytes):
    try:
        return load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        return None


def _find_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower().startswith("publikime al"):
            return wb[name]
    return None


def _extract_date(ws):
    value = ws[CELL_DATE].value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%d.%m.%Y").date()
        except ValueError:
            return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value)).date()
    return None


def _extract_rows(ws, rep_date):
    vals = [ws[f"F{row}"].value for row in range(ROW_START, ROW_END)]
    return [
        {"date": rep_date.isoformat(), "hour": h + 1, "demand": v}
        for h, v in enumerate(vals) if v is not None
    ]

# ─────────────────────────────────────────────────────────────────────────── #

def run(overwrite=False):
    # 1. config ------------------------------------------------------------- #
    cfg_path = os.path.join(os.path.dirname(__file__), "..", "config.yaml")
    with open(cfg_path, encoding="utf-8") as fp:
        cfg = yaml.safe_load(fp)

    start = datetime.strptime(cfg["START_DATE"], "%Y-%m-%d")
    end   = datetime.strptime(cfg["END_DATE"], "%Y-%m-%d")
    out_dir = cfg["OUTPUT_DIR"]
    max_workers = min(10, int(cfg.get("MAX_WORKERS", DEFAULT_WORKERS)))
    os.makedirs(out_dir, exist_ok=True)

    wanted_days = pd.date_range(start, end).date
    search_days = pd.date_range(start, end + relativedelta(months=1)).date

    # 2. worker ------------------------------------------------------------- #
    def fetch(day):
        sess = _session()
        local_success = None  # remember the first working pair to reorder loop

        # Re‑order suffixes: put any previously successful suffixes first
        ordered_suffixes = sorted(SUFFIXES, key=lambda s: (0 if (day.year, day.month, s) in working else 1))

        for suf in ordered_suffixes:
            for off in FOLDERS:
                folder = datetime(day.year, day.month, 1) + relativedelta(months=off)
                key = (folder.year, folder.month, suf)

                with _cache_lock:
                    if key in fail_count and fail_count[key] >= FAIL_THRESHOLD and key not in working:
                        continue

                url = (
                    f"{BASE_URL}/{folder.year}/{folder.month:02d}/"
                    f"Publikimi-te-dhenave-{day.day:02d}.{day.month:02d}.{day.year}{suf}.xlsx"
                )
                try:
                    do_full_get = True
                    if key not in working:
                        # try HEAD first
                        h = sess.head(url, timeout=HTTP_TIMEOUT, allow_redirects=True)
                        if h.status_code == 404:
                            do_full_get = False  # still might GET once below
                    if do_full_get:
                        resp = sess.get(url, timeout=HTTP_TIMEOUT)
                        if resp.status_code != 200:
                            raise requests.RequestException("non‑200")
                        wb = _open_wb(resp.content)
                        if wb is None:
                            raise requests.RequestException("bad wb")
                        ws = _find_sheet(wb)
                        if ws is None:
                            raise requests.RequestException("sheet")
                        rep_date = _extract_date(ws)
                        if not rep_date or not (start.date() <= rep_date <= end.date()):
                            raise requests.RequestException("date")
                        rows = _extract_rows(ws, rep_date)
                        if rows:
                            with _cache_lock:
                                working.add(key)
                            return rep_date, rows
                except requests.RequestException:
                    with _cache_lock:
                        fail_count[key] = fail_count.get(key, 0) + 1
                    continue
        return None

    # 3. concurrent fetch --------------------------------------------------- #
    bar = tqdm(total=len(search_days), desc="OST", unit="file", disable=VERBOSE, dynamic_ncols=True)
    collected = {}

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(fetch, d): d for d in search_days}
        try:
            for fut in as_completed(futures):
                bar.update(1)
                res = fut.result()
                if res:
                    rep, rows = res
                    if rep not in collected or len(rows) > len(collected[rep]):
                        collected[rep] = rows
                    if len(collected) == len(wanted_days):  # all days covered
                        for f in futures:
                            f.cancel()
                        break
        except CancelledError:
            pass
        finally:
            bar.close()

    # 4. skeleton + merge --------------------------------------------------- #
    skel = pd.MultiIndex.from_product([
        pd.Series(wanted_days).astype(str), range(1, 25)
    ], names=["date", "hour"]).to_frame(index=False)

    real = pd.DataFrame([r for lst in collected.values() for r in lst],
                        columns=["date", "hour", "demand"])

    df = skel.merge(real, how="left", on=["date", "hour"]).sort_values(["date", "hour"])

    # 5. output ------------------------------------------------------------- #
    df["datetime"] = pd.to_datetime(df["date"]) + pd.to_timedelta(df["hour"] - 1, unit="h")
    final_df = df[["datetime", "demand"]].sort_values("datetime")

    out_csv = os.path.join(out_dir, "ost_demand.csv")
    final_df.to_csv(out_csv, index=False, na_rep="")
    print(
        f"✅ Saved {len(final_df):,} rows "
        f"({final_df['datetime'].dt.date.nunique()} day(s)) → {out_csv}"
    )

    missing = sorted(set(wanted_days) - set(final_df["datetime"].dt.date))
    if missing:
        print("⚠ No workbook found for:", ", ".join(d.isoformat() for d in missing))
